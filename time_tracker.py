import datetime
from datetime import timedelta
import datetime as dt
from openpyxl.styles import Font
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Fill
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.cell import column_index_from_string
import os

#################################################################################################
# Configurations
#################################################################################################
ROOT_PATH = r"C:\Users\m0pxnn\Documents\TimeTracking"
HEADING_FIELDS = ["     Date    ", "   Day   ", "   InTime   ", "   OutTime   ", "   Hours   "]
NUM_FIELDS = len(HEADING_FIELDS)
DATA_START_ROW = 1
MAX_MONTH_DAYS = 30
MAX_ROWS = MAX_MONTH_DAYS + 1
MAX_COLS = NUM_FIELDS
VERBOSE_OUTPUT = False
REQUIRED_HOURS = [8, 30] # 8 hrs 30 minutes

# Index of different columns in excel sheet
DATE_INDEX    = 0
WEEKDAY_INDEX = 1
INTIME_INDEX  = 2
OUTTIME_INDEX = 3
HOURS_INDEX   = 4

AVG_INTIME_HEADING  = " Avg InTime "
AVG_OUTTIME_HEADING = " Avg OutTime "
AVG_HOURS_HEADING   = " Avg hours "
REQ_HOURS_HEADING   = " Req hours "
TOTAL_HOURS_HEADING = " Total hours "

COL_DATE = 'A'
COL_DAY = 'B'
COL_INTIME = 'C'
COL_OUTTIME = 'D'
COL_HOURS = 'E'
COL_REQ_HOURS = 'F'
COL_TOTAL_HOURS = 'G'
COL_DIFF_MSG = 'H'

CELL_AVG_INTIME_HEADING  = COL_INTIME + str(MAX_ROWS)
CELL_AVG_OUTTIME_HEADING = COL_OUTTIME + str(MAX_ROWS)
CELL_AVG_HOURS_HEADING   = COL_HOURS + str(MAX_ROWS)
CELL_REQ_HOURS_HEADING   = COL_REQ_HOURS + str(MAX_ROWS)
CELL_TOTAL_HOURS_HEADING = COL_TOTAL_HOURS + str(MAX_ROWS)


CELL_AVG_INTIME_DATA  = COL_INTIME + str(MAX_ROWS + 1)
CELL_AVG_OUTTIME_DATA = COL_OUTTIME + str(MAX_ROWS + 1)
CELL_AVG_HOURS_DATA   = COL_HOURS + str(MAX_ROWS + 1)
CELL_REQ_HOURS_DATA   = COL_REQ_HOURS + str(MAX_ROWS + 1)
CELL_TOTAL_HOURS_DATA = COL_TOTAL_HOURS + str(MAX_ROWS + 1)
CELL_DIFF_MSG_DATA    = COL_DIFF_MSG + str(MAX_ROWS + 1)



#################################################################################################
# Functions
#################################################################################################

#################################################################################################
# @name         : WriteMiscValuesToSheet
# @description  : At the end of the sheet, write down the Average intime, outtime and hours.
#################################################################################################
def WriteMiscValuesToSheet(sheet, avg_inTime, avg_outTime, avg_hours, requiredHours, totalHours, diffMessage):
    # Add font to Average row
    avgRowFont = Font(color='00000000', bold=True)
       
    # Write misc. details to sheet
    sheet.column_dimensions[COL_INTIME].width = len(AVG_INTIME_HEADING)
    sheet[CELL_AVG_INTIME_HEADING].font = avgRowFont
    sheet[CELL_AVG_INTIME_HEADING] = AVG_INTIME_HEADING
    sheet[CELL_AVG_INTIME_DATA] = avg_inTime
    
    sheet.column_dimensions[COL_OUTTIME].width = len(AVG_OUTTIME_HEADING)
    sheet[CELL_AVG_OUTTIME_HEADING].font = avgRowFont
    sheet[CELL_AVG_OUTTIME_HEADING] = AVG_OUTTIME_HEADING
    sheet[CELL_AVG_OUTTIME_DATA] = avg_outTime
    
    sheet.column_dimensions[COL_HOURS].width = len(AVG_HOURS_HEADING)
    sheet[CELL_AVG_HOURS_HEADING].font = avgRowFont
    sheet[CELL_AVG_HOURS_HEADING] = AVG_HOURS_HEADING
    sheet[CELL_AVG_HOURS_DATA] = avg_hours
    
    sheet.column_dimensions[COL_REQ_HOURS].width = len(REQ_HOURS_HEADING)
    sheet[CELL_REQ_HOURS_HEADING].font = avgRowFont
    sheet[CELL_REQ_HOURS_HEADING] = REQ_HOURS_HEADING
    sheet[CELL_REQ_HOURS_DATA] = float('%.2f'%(requiredHours))
    
    
    sheet.column_dimensions[COL_TOTAL_HOURS].width = len(TOTAL_HOURS_HEADING)
    sheet[CELL_TOTAL_HOURS_HEADING].font = avgRowFont
    sheet[CELL_TOTAL_HOURS_HEADING] = TOTAL_HOURS_HEADING
    sheet[CELL_TOTAL_HOURS_DATA] = float('%.2f'%(totalHours))
    
       
    sheet.column_dimensions[COL_DIFF_MSG].width = len(diffMessage)
    sheet[CELL_DIFF_MSG_DATA] = diffMessage
    
    
#################################################################################################
# @name         : CreateNewWorkbook
# @description  : Creates a new excel document corresponding to this month. Adds heading row 
#                 and formatting for it and saves it.
#################################################################################################
def CreateNewWorkbook(fileNameWithPath):
    book = Workbook()
    sheet = book.active    
    
    row_heading = HEADING_FIELDS
    sheet.append(row_heading)
    
    # Adjust width of each column
    i = 0
    while (i < len(HEADING_FIELDS)):
        sheet.column_dimensions[get_column_letter(i+1)].width = len(HEADING_FIELDS[i])
        #sheet[get_column_letter(i+1)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        i = i + 1
    
    # Add font to heading row
    headingRowFont = Font(color='00000000', bold=True)
    for cell in sheet["1:1"]:
        cell.font = headingRowFont
    
    # Save the workbook
    book.save(fileNameWithPath)
    book.close()
    

#################################################################################################
# @name         : PrepareDataForToday
# @description  : Opens the excel sheet for this month and looks for presence of today date. If
#                 available, it updates its outTime and Hours fields. If not available, it will
#                 add a new entry for current day.
#################################################################################################    
def PrepareDataForToday(fileNameWithPath, dateTimeObj):
    book = openpyxl.load_workbook(fileNameWithPath)
    sheet = book.active 
    
    date = dateTimeObj.strftime("%d-%b-%Y")
    time = dateTimeObj.strftime("%H:%M:%S")   
    weekDay = dateTimeObj.strftime("%a")
    
    inTime = time
    outTime = ""
    hours = ""
    
    totalEntries = 0
    recordFound = False
    rowsFilled = sheet.max_row
    
    seconds_inTime_total = None;
    seconds_outTime_total = None;
    seconds_total = None;   
    
    rowNumber = 0
    
    for row in sheet.rows:   
        rowNumber = rowNumber + 1
        
        # Date value in this row of the sheet
        sheet_date = row[DATE_INDEX].value
        if (sheet_date == HEADING_FIELDS[DATE_INDEX]):
            # This is the heading row, skip it
            continue
        
        if (rowNumber >= MAX_ROWS or sheet_date == None):
            break
        
        # Week day in this row of the sheet    
        sheet_weekDay = row[WEEKDAY_INDEX].value
        
        # InTime value in this row of the sheet. This field should NOT BE 'None', because
        # if this entry is present, it must have an inTime.
        if (VERBOSE_OUTPUT):
            print()
            print (sheet_date)
            
        sheet_inTime = row[INTIME_INDEX].value
        if (sheet_inTime != None):            
            sheet_inTime_dt = datetime.datetime.strptime(sheet_inTime, "%H:%M:%S")
            tmp = datetime.timedelta(hours=sheet_inTime_dt.hour, 
                                     minutes=sheet_inTime_dt.minute, 
                                     seconds=sheet_inTime_dt.second).total_seconds()
                                     
            if (seconds_inTime_total == None):
                seconds_inTime_total = tmp              
            else:
                seconds_inTime_total = seconds_inTime_total + tmp 
            
            if (VERBOSE_OUTPUT):
                print ("sheet_inTime  : ", sheet_inTime)
                print ("  seconds     : ", seconds_inTime_total)
            
        # OutTime value in this row of the sheet. This field CAN be 'None'.       
        sheet_outTime = row[OUTTIME_INDEX].value
        if (sheet_outTime != None):            
            sheet_outTime_dt = datetime.datetime.strptime(sheet_outTime, "%H:%M:%S")
            tmp = datetime.timedelta(hours=sheet_outTime_dt.hour, 
                                     minutes=sheet_outTime_dt.minute, 
                                     seconds=sheet_outTime_dt.second).total_seconds()
                                     
            if (seconds_outTime_total == None):
                seconds_outTime_total = tmp              
            else:
                seconds_outTime_total = seconds_outTime_total + tmp    
            
            if (VERBOSE_OUTPUT):
                print ("sheet_outTime : ", sheet_outTime)            
                print ("  seconds     : ", seconds_outTime_total)
        
        # Hours value in this row of the sheet. This CAN be 'None'
        sheet_hours = row[HOURS_INDEX].value
        if (sheet_hours != None):
            tmp = datetime.timedelta(hours=sheet_hours.hour, 
                                     minutes=sheet_hours.minute, 
                                     seconds=sheet_hours.second).total_seconds()
                                     
            if (seconds_total == None):
                seconds_total = tmp              
            else:
                seconds_total = seconds_total + tmp               
            
            if (VERBOSE_OUTPUT):
                print ("  seconds     : ", seconds_total)
                print ("sheet_hours   : ", sheet_hours)
                
        # If this record has the same date as the current date, then we need to update
        # the outTime of this entry and re-calculate the Hours.
        if (date == sheet_date):
            recordFound = True
            inTime = sheet_inTime
            outTime = time                             
                
            #hours = outTime - inTime
            hours = datetime.datetime.strptime(outTime, "%H:%M:%S") - datetime.datetime.strptime(inTime, "%H:%M:%S")
            
            # Update in excel sheet
            print()
            print("Updating entry...")          
            row[OUTTIME_INDEX].value = outTime
            row[HOURS_INDEX].value = hours   

            # Since we are breaking the loop, update the entry count.
            totalEntries = totalEntries + 1
            break


        # Update the entries present in the sheet
        totalEntries = totalEntries + 1   
       
    
    print("Total entries found: ", totalEntries)
    
    # If entry for this date is not present, only then we need to add this entry, else
    # we need to just update the current record.
    if (not recordFound):
        print ()
        print("Adding entry...")          
        newRow = [date, weekDay, inTime, outTime, hours]
        #sheet.append(newRow)

        sheet.cell(row=totalEntries+1, column=column_index_from_string(COL_DATE)).value = date
        sheet.cell(row=totalEntries+1, column=column_index_from_string(COL_DAY)).value = weekDay
        sheet.cell(row=totalEntries+1, column=column_index_from_string(COL_INTIME)).value = inTime
        sheet.cell(row=totalEntries+1, column=column_index_from_string(COL_OUTTIME)).value = outTime
        sheet.cell(row=totalEntries+1, column=column_index_from_string(COL_HOURS)).value = hours
        
    
    # Calculate expected outTime for today (as per the REQUIRED_HOURS parameter)
    secondsRequiredToday = (REQUIRED_HOURS[0] * 60 * 60) + (REQUIRED_HOURS[1] * 60)
    expectedOutTimeForToday = datetime.timedelta(seconds=secondsRequiredToday) + datetime.datetime.strptime(inTime, "%H:%M:%S")
    expectedOutTimeForToday = expectedOutTimeForToday.strftime("%H:%M:%S")    
    
    if (totalEntries > 1):
        # Update average info only if we have 2 or more entries
        avg_inTime_seconds = seconds_inTime_total / totalEntries
        avg_outTime_seconds = seconds_outTime_total / totalEntries
        avg_seconds = seconds_total / totalEntries  
        
        
        avg_inTime = datetime.timedelta(seconds=avg_inTime_seconds)
        avg_outTime = datetime.timedelta(seconds=avg_outTime_seconds)
        avg_hours = datetime.timedelta(seconds=avg_seconds)
        
        requiredSeconds = (totalEntries * REQUIRED_HOURS[0] * 60 * 60) + (totalEntries * REQUIRED_HOURS[1] * 60)
        requiredHours = requiredSeconds / 60 / 60
        totalHours    = seconds_total / 60 / 60
        diffHours     = requiredHours - totalHours
        diffHours     = datetime.timedelta(hours=diffHours) 
        
        if (seconds_total >= requiredSeconds):
            diffMessage = " (Target Met) "
        else:
            diffMessage = " (Missing target) "
        
    # Display this info on console
    print()
    print("** Today **")
    print("Date              : ", date)
    print("Day               : ", weekDay)
    print("In Time           : ", inTime)
    print("Required Out Time : ", expectedOutTimeForToday)
    print("Actual Out Time   : ", outTime)
    print("Hours             : ", hours)
    #print("---------------------------------------")
    
    if (totalEntries > 1):
        # If multiple records were found, only then we need to calculate avg values
        print()
        print("** This month **")
        print("Required hours    :  %s hours" % '%.2f'%(requiredHours))
        print("Total hours       :  %s hours" % '%.2f'%(totalHours))
        print("Difference        :  {} {}".format(diffHours, diffMessage))
        print("Average In Time   : ", avg_inTime)
        print("Average Out Time  : ", avg_outTime)
        print("Average Hours     : ", avg_hours)  

        # Write the average hours, in time & out time values at the end of the sheet
        WriteMiscValuesToSheet(sheet, avg_inTime, avg_outTime, avg_hours, requiredHours, totalHours, diffMessage)
      
    # Save the excel sheet
    print("\nSaving [%s]\n" % fileNameWithPath)
    book.save(fileNameWithPath)
    book.close()
    
#################################################################################################
# Main
#################################################################################################

print()
print("+---------------------------------------------------------------------+")
print("|                      T I M E     T R A C K E R                      |")
print("+---------------------------------------------------------------------+")

# Current date-time. This will determine the entries in the time tracker.
dateTimeObj = datetime.datetime.now()

# Find filename of Excel that should hold this record
fileName = dateTimeObj.strftime("%b-%Y")
fileName = fileName + ".xlsx"
fileNameWithPath = os.path.join(ROOT_PATH, fileName)

# Open and Read excel for this month
print("Checking [%s]..." % fileNameWithPath)
if (not os.path.isfile(fileNameWithPath)):
    print("Creating workbook for this month...")
    CreateNewWorkbook(fileNameWithPath)

row_data = PrepareDataForToday(fileNameWithPath, dateTimeObj)

print("+---------------------------------------------------------------------+")
